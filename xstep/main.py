from openpyxl import load_workbook

from xstep.testing import Testing
from tools.visa_cmd_rs import *

if __name__ == '__main__':
    # t = Testing()
    # rf_params = {"freq": "2190", "bandwidth": "20", "mode": "31a",
    #              "branch": "7", "loss": "28.6", "power": "23", "temp": "40"}
    # t.measure(rf_params)

    # v = VisaConnection("TCPIP0::192.168.0.130::inst0::INSTR", timeout=10000)
    # config_list = Testing.getConfig("./config/4G_config.xml")
    # wb = load_workbook("./report/common.xlsx")
    # ws_sem = wb["SEM"]
    # ws_se = wb["SE"]
    # row_sem = 26
    # row_se = 32
    # for config in config_list:
    #     config['temp'] = '25'
    #     config['mode'] = '31'
    #     input("branch -> " + config['branch'] + "; bandwidth -> " + config['bandwidth'] +
    #           "; frequency -> " + config['freq'] + " Press Enter to continue!")
    #
    #     v.send_cmd("*RST")
    #     v.send_cmd("INST:CRE:NEW LTE, 'LTE'")
    #     res2 = lte_sem(v, config, current='LTE', rename='T1', snap=True, delay=10, atten=15,
    #                    snappath=r"C:\\Xstep\\SEM\\%s_%sM_%s_1.JPG" % (config['branch'], config['bandwidth'], config['freq']))
    #     res1 = [int(config['bandwidth']), float(config['freq']), int(config['branch']), float(config['power']), res2[0]]
    #     for r in ws_sem.iter_rows(min_row=row_sem, max_row=row_sem, max_col=5):
    #         for cell in r:
    #             cell.value = res1[cell.column-1]
    #
    #     i = 1
    #     for r in ws_sem.iter_rows(min_row=row_sem, max_row=row_sem+3, min_col=6, max_col=13):
    #         for cell in r:
    #             cell.value = res2[i]
    #             i += 1
    #     row_sem = row_sem + 4
    #
    #
    #     v.send_cmd("INST:CRE:NEW SANALYZER, 'Spectrum'")
    #     res3 = se(v, config, current='Spectrum', rename='T2', delay=10, snap=True,
    #               snappath=r"C:\\Xstep\\SEM\\%s_%sM_%s_2.JPG" % (config['branch'], config['bandwidth'], config['freq']))
    #     for r in ws_se.iter_rows(min_row=row_se, max_row=row_se, max_col=5):
    #         for cell in r:
    #             cell.value = res1[cell.column-1]
    #     j = 0
    #     for r in ws_se.iter_rows(min_row=row_se, max_row=row_se+4, min_col=5, max_col=11):
    #         for cell in r:
    #             cell.value = res3[j]
    #             j += 1
    #     row_se = row_se + 5
    #
    #     wb.save("./report/common.xlsx")



    v = VisaConnection("TCPIP0::192.168.0.130::inst0::INSTR", timeout=1000000)
    config_list = Testing.getConfig("./config/5G_config.xml")
    wb = load_workbook("./report/common.xlsx")
    ws_sem = wb["SEM"]
    ws_se = wb["SE"]
    row_sem = 98
    row_se = 127
    for config in config_list:
        config['temp'] = '25'
        config['mode'] = '31'
        input("branch -> " + config['branch'] + "; bandwidth -> " + config['bandwidth'] +
              "; frequency -> " + config['freq'] + " Press Enter to continue!")

        v.send_cmd("*RST")
        v.send_cmd("INST:CRE:NEW NR5G, '5G NR'")
        res2 = nr5g_sem(v, config, current='5G NR', rename='T3', delay=30, atten=15, exs=True, snap=True,
                       snappath=r"C:\\Xstep\\SEM\\%s_%sM_%s_1.JPG" % (config['branch'], config['bandwidth'], config['freq']))
        res1 = [int(config['bandwidth']), float(config['freq']), int(config['branch']), float(config['power']), res2[0]]
        for r in ws_sem.iter_rows(min_row=row_sem, max_row=row_sem, max_col=5):
            for cell in r:
                cell.value = res1[cell.column-1]

        i = 1
        for r in ws_sem.iter_rows(min_row=row_sem, max_row=row_sem+3, min_col=6, max_col=13):
            for cell in r:
                cell.value = res2[i]
                i += 1
        row_sem = row_sem + 4


        v.send_cmd("INST:CRE:NEW SANALYZER, 'Spectrum'")
        res3 = se(v, config, current='Spectrum', rename='T4', delay=10, snap=True,
                  snappath=r"C:\\Xstep\\SEM\\%s_%sM_%s_2.JPG" % (config['branch'], config['bandwidth'], config['freq']))
        for r in ws_se.iter_rows(min_row=row_se, max_row=row_se, max_col=5):
            for cell in r:
                cell.value = res1[cell.column-1]
        j = 0
        for r in ws_se.iter_rows(min_row=row_se, max_row=row_se+4, min_col=5, max_col=11):
            for cell in r:
                cell.value = res3[j]
                j += 1
        row_se = row_se + 5

        wb.save("./report/common.xlsx")