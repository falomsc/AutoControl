import configparser

from lxml import etree
from openpyxl import load_workbook

from tools.soaptalker import SoapTalker
from tools.ssh import SSHConnection
from tools.ssh_cmd import get_atten
from tools.visa_cmd_rs import *

file_dict = {"0": "N78", "1": "N78", "2": "N78", "3": "N78", "4": "B7", "5": "B7", "6": "B66", "7": "B66"}

cf = configparser.ConfigParser()
cf.read("./config/equip_config.ini")


class Testing:
    def __init__(self):
        """
        s1: establish exit trigger
        s2: get attenuation
        :param pid:
        """
        self.s1 = SSHConnection(hostname=cf.get("ssh", "hostname"), port=int(cf.get("ssh", "port")),
                                username=cf.get("ssh", "username"), password=cf.get("ssh", "password"),
                                print_flag=False)
        self.s2 = SSHConnection(hostname=cf.get("ssh", "hostname"), port=int(cf.get("ssh", "port")),
                                username=cf.get("ssh", "username"), password=cf.get("ssh", "password"),
                                print_flag=False)
        self.v = VisaConnection(address=cf.get("visa", "address"), timeout=int(cf.get("visa", "timeout")))
        # self.so = SoapTalker(cf.get("soaptalker", "path"))

    def get_acp(self, rf_params: Dict[str, str], ws: Worksheet, row: int):
        if rf_params.get("pipe") == 0:
            self.v.send_cmd("INST:CRE:NEW NR5G, '5G NR'")
            res = nr5g_aclr(self.v, rf_params, current='5G NR', rename='T1', exs=True, delay=20)
        else:
            self.v.send_cmd("INST:CRE:NEW LTE, 'LTE'")
            res = lte_aclr(self.v, rf_params, current='LTE', rename='T1')
        ##############################
        res_acp = [res[0], res[3], res[1], res[2], res[4],
                   res[1] if res[1] > res[2] else res[2], res[3] if res[3] > res[4] else res[4]
                   ]
        for r in ws.iter_rows(min_row=row, max_row=row, min_col=9, max_col=15):
            for cell in r:
                cell.value = res_acp[cell.column - 9]

    def get_obw(self, rf_params: Dict[str, str], ws: Worksheet, row: int):
        self.v.send_cmd("INST:CRE:NEW SANALYZER, 'Spectrum'")
        span = 30 if rf_params.get("bandwidth") == "20" else 120
        res = obw(self.v, rf_params, span, current='Spectrum', rename='T2')
        ws.cell(row=row, column=16).value = res

    def get_evm(self, rf_params: Dict[str, str], ws: Worksheet, row: int):
        if self.check5G(rf_params):
            self.v.send_cmd("INST:CRE:NEW NR5G, '5G NR'")
            res = nr5g_evm(self.v, rf_params, current='5G NR', rename='T3', exs=True, delay=10)
        else:
            self.v.send_cmd("INST:CRE:NEW LTE, 'LTE'")
            res = lte_evm(self.v, rf_params, current='LTE', rename='T3')
        # 64QAM TM31 TM2
        if rf_params.get('mode') in ("2", "31"):
            res_evm = [res[2][0], res[7][0], res[12 if self.check5G(rf_params) else 13][0]]
        # 256QAM TM31a TM2a
        else:
            res_evm = [res[3][0], res[7][0], res[12 if self.check5G(rf_params) else 13][0]]
        col_start = 17 if rf_params.get('mode') in ("31", "31a") else 7
        for r in ws.iter_rows(min_row=row, max_row=row, min_col=col_start, max_col=col_start + 2):
            for cell in r:
                cell.value = res_evm[cell.column - col_start]

    def get_ccdf(self, rf_params: Dict[str, str], ws: Worksheet, row: int):
        self.v.send_cmd("INST:CRE:NEW SANALYZER, 'Spectrum'")
        abw = 80 if rf_params.get("bandwidth") == "20" else 240
        res = ccdf(self.v, rf_params, abw, current='Spectrum', rename='T4')
        ws.cell(row=row, column=20).value = res[-1]

    def obtain_atten(self, rf_params: Dict[str, str], ws: Worksheet, row: int):
        atten = get_atten(self.s2, rf_params)
        ws.cell(row=row, column=6).value = atten

    def measure(self, rf_params: Dict[str, str]):
        wb = load_workbook("./report/%s.xlsx" % file_dict[rf_params.get('branch')])
        ws = wb["TM%s_%s" % (rf_params.get('mode'), rf_params.get('bandwidth'))]
        row = newline(ws, init_row=3)

        # basic information
        res_basic = [int(rf_params.get('temp')), int(rf_params.get('bandwidth')), float(rf_params.get('freq')),
                     int(rf_params.get('branch')), float(rf_params.get('power'))]
        for r in ws.iter_rows(min_row=row, max_row=row, max_col=5):
            for cell in r:
                cell.value = res_basic[cell.column - 1]

        # FSW Test
        self.v.send_cmd("*RST")
        self.v.send_cmd("*CLS")
        if rf_params.get("mode") in ("31", "31a"):
            self.get_acp(rf_params, ws, row)
            self.get_obw(rf_params, ws, row)
            self.get_evm(rf_params, ws, row)
            self.get_ccdf(rf_params, ws, row)
            self.obtain_atten(rf_params, ws, row)
        elif rf_params.get("mode") in ("2", "2a"):
            self.get_evm(rf_params, ws, row)
            self.obtain_atten(rf_params, ws, row)

        wb.save("./report/%s.xlsx" % file_dict[rf_params.get('branch')])

    def test_process(self, rf_params: Dict[str, str]):
        self.so.message_init()
        self.so.tx_on(tx_type=int(rf_params.get('branch')) % 4 + 1)
        self.measure(rf_params)
