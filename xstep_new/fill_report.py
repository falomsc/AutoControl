from typing import Union

from openpyxl import load_workbook

from tools.function import new_line


class FillReport:
    BASIC = 10
    MADURA = 22
    LTE_ACP = 30
    LTE_MULTI_ACP = 31
    LTE_MULTI_GAP_ACP = 32
    NR5G_ACP = 33
    NR5G_MULTI_ACP = 34
    NR5G_MULTI_GAP_ACP = 35
    OBW = 40
    MULTI_OBW = 41
    CCDF = 50
    LTE_EVM = 60
    LTE_MULTI_EVM = 61
    NR5G_EVM = 62
    NR5G_MULTI_EVM = 63
    LTE_SEM = 70
    LTE_MULTI_SEM = 71
    LTE_MULTI_GAP_SEM = 72
    NR5G_SEM = 73
    NR5G_MULTI_SEM = 74
    NR5G_MULTI_GAP_SEM = 75
    SE = 80

    def __init__(self, path: str):
        self._path = path
        self._wb = load_workbook(self._path)
        self._ws = self._wb["RAW_DATA"]
        self._first_row = tuple(self._ws.values)[0]
        self._second_row = tuple(self._ws.values)[1]
        self._row = 3

    def next_line(self):
        self._row = new_line(self._ws, init_row=self._row, col=3)

    def fill_xl_value(self, typ: int, res: Union[list, float], mode_extra: str = "TM3_1"):
        col_start = 1

        if typ == FillReport.BASIC:
            for i, v in enumerate(self._second_row):
                if v == "Test_Mode":
                    col_start = i + 1
                    break
            for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + len(res) - 1):
                for j, cell in enumerate(r):
                    cell.value = res[j]

        if typ == FillReport.MADURA:
            for i, v in enumerate(self._second_row):
                if v == "Madura_Atten(dB)":
                    col_start = i + 1
                    break
            for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + len(res) - 1):
                for j, cell in enumerate(r):
                    cell.value = res[j]

        if typ // 10 == 3:
            for i, v in enumerate(self._second_row):
                if v == "CHP1(dBm)":
                    col_start = i + 1
                    break
            if typ in (FillReport.LTE_ACP, FillReport.NR5G_ACP):
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start + 3, max_col=col_start + 7):
                    for j, cell in enumerate(r):
                        cell.value = res[j]
            if typ == FillReport.LTE_MULTI_ACP:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 7):
                    res.insert(3, "")
                    for j, cell in enumerate(r):
                        cell.value = res[j]
            if typ == FillReport.NR5G_MULTI_ACP:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 7):
                    for j, cell in enumerate(r):
                        cell.value = res[j]
            if typ in (FillReport.LTE_MULTI_GAP_ACP, FillReport.NR5G_MULTI_GAP_ACP):
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 11):
                    for j, cell in enumerate(r):
                        cell.value = res[j]

        elif typ // 10 == 4:
            for i, v in enumerate(self._second_row):
                if v == "OBW1(MHz)":
                    col_start = i + 1
                    break
            if typ == FillReport.OBW:
                self._ws.cell(row=self._row, column=col_start).value = res
            if typ == FillReport.MULTI_OBW:
                self._ws.cell(row=self._row, column=col_start).value = res[0]
                self._ws.cell(row=self._row, column=col_start + 1).value = res[1]

        elif typ == FillReport.CCDF:
            for i, v in enumerate(self._second_row):
                if v == "CCDF(dB)":
                    col_start = i + 1
                    break
            self._ws.cell(row=self._row, column=col_start).value = res[6]

        elif typ // 10 == 6:
            for i, v in enumerate(self._second_row):
                if v == "EVM1(%)":
                    col_start = i + 1
                    break
            if typ == FillReport.LTE_EVM:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 11):
                    if mode_extra in ("TM3_1", "TM2"):
                        r[0].value = res[2][0]
                    elif mode_extra in ("TM3_1A", "TM2A"):
                        r[0].value = res[3][0]
                    r[2].value = res[7][0]
                    r[4].value = res[13][0]
            elif typ == FillReport.LTE_MULTI_ACP:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 11):
                    if mode_extra in ("TM3_1", "TM2"):
                        r[0].value = res[0][2][0]
                        r[1].value = res[1][2][0]
                    elif mode_extra in ("TM3_1A", "TM2A"):
                        r[0].value = res[0][3][0]
                        r[1].value = res[1][3][0]
                    r[2].value = res[0][7][0]
                    r[3].value = res[1][7][0]
                    r[4].value = res[0][13][0]
                    r[5].value = res[0][13][0]
            elif typ == FillReport.NR5G_EVM:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 11):
                    if mode_extra in ("TM3_1", "TM2"):
                        r[0].value = res[2][0]
                    elif mode_extra in ("TM3_1A", "TM2A"):
                        r[0].value = res[3][0]
                    r[2].value = res[7][0]
                    r[4].value = res[12][0]
            elif typ == FillReport.NR5G_MULTI_EVM:
                for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + 11):
                    if mode_extra in ("TM3_1", "TM2"):
                        r[0].value = res[0][2][0]
                        r[1].value = res[1][2][0]
                    elif mode_extra in ("TM3_1A", "TM2A"):
                        r[0].value = res[0][3][0]
                        r[1].value = res[1][3][0]
                    r[2].value = res[0][7][0]
                    r[3].value = res[1][7][0]
                    r[4].value = res[0][12][0]
                    r[5].value = res[0][12][0]

        elif typ // 10 == 7:
            if typ in (FillReport.LTE_SEM, FillReport.NR5G_SEM):
                n = 1
            elif typ in (FillReport.LTE_MULTI_SEM, FillReport.NR5G_MULTI_SEM):
                n = 2
            for i, v in enumerate(self._second_row):
                if v == "Range No.1" and self._ws.cell(row=1, column=i + 1).value == "SEM":
                    col_start = i + 1
                    break
            for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start,
                                        max_col=col_start + len(res) - n - 1):
                for j, cell in enumerate(r):
                    cell.value = res[j + n]

        elif typ == FillReport.SE:
            for i, v in enumerate(self._second_row):
                if v == "Range No.1" and self._ws.cell(row=1, column=i + 1).value == "SE":
                    col_start = i + 1
                    break
            for r in self._ws.iter_rows(min_row=self._row, max_row=self._row, min_col=col_start, max_col=col_start + len(res) - 1):
                for j, cell in enumerate(r):
                    cell.value = res[j]

        self._wb.save(self._path)


if __name__ == '__main__':
    fil = FillReport("./report/model.xlsx")
    # fil.fill_xl_value(FillReport.LTE_MULTI_ACP, 5, [20, 20, 23, -45, -46, -47, -48])
